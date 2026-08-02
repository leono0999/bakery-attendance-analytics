import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def generate_employee_reports(attendance_path, templates_folder, reports_folder):
    """Generate one Excel report per employee by copying their template and
    filling in the CALCULO sheet with this period's attendance rows.

    Existing formulas/formatting in each template are preserved; only the
    date, entrada, and salida cells are overwritten.
    """
    attendance_path = Path(attendance_path)
    templates_folder = Path(templates_folder)
    reports_folder = Path(reports_folder)
    reports_folder.mkdir(exist_ok=True)

    # Clear out reports from a previous run
    for file in reports_folder.glob("*.xlsx"):
        file.unlink()

    df = pd.read_excel(attendance_path)
    df["FirstName"] = df["Nombre"].str.split().str[0].str.upper()

    for employee, group in df.groupby("FirstName"):
        employee_file = next(
            (
                file
                for file in templates_folder.iterdir()
                if file.name.upper().startswith(employee)
                and file.suffix.lower() == ".xlsx"
            ),
            None,
        )
        if employee_file is None:
            print(f"File not found for {employee}")
            continue

        output_path = reports_folder / employee_file.name
        print(f"\nProcessing {employee_file.name}")
        shutil.copy2(employee_file, output_path)

        wb = load_workbook(output_path)
        calculo_sheet = next(
            (s for s in wb.sheetnames if s.strip().upper() == "CALCULO"), None
        )
        if calculo_sheet is None:
            print(f"CALCULO sheet not found for {employee}")
            continue
        ws = wb[calculo_sheet]

        # Clear old data rows before writing this period's data
        for row in range(8, 45):
            ws[f"B{row}"] = None
            ws[f"E{row}"] = None
            ws[f"F{row}"] = None

        group = group.sort_values("Fecha")
        start_row = 8
        for i, (_, attendance_row) in enumerate(group.iterrows()):
            row = start_row + i
            attendance_date = pd.to_datetime(attendance_row["Fecha"])
            ws[f"B{row}"] = attendance_date
            ws[f"B{row}"].number_format = "dd/mm/yyyy"

            entrada = pd.to_datetime(attendance_row["Entrada"])
            salida = pd.to_datetime(attendance_row["Salida"])

            # Excel time-only cells use the 1899-12-30 epoch
            ws[f"E{row}"] = datetime(
                1899, 12, 30, entrada.hour, entrada.minute, entrada.second
            )
            ws[f"F{row}"] = datetime(
                1899, 12, 30, salida.hour, salida.minute, salida.second
            )
            ws[f"E{row}"].number_format = "hh:mm:ss"
            ws[f"F{row}"].number_format = "hh:mm:ss"

        wb.save(output_path)
        print(f"Created: {output_path}")

    print("\nDone!")
